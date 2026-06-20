from fastapi import APIRouter, Depends, Header
from fastapi.responses import StreamingResponse
from sqlalchemy.orm import Session
from sqlalchemy import func as sqlfunc, or_
from datetime import datetime, date
from typing import Optional, List
import json
import io

from database import get_db
from models import (
    Producto, VarianteProducto, VendedorPerfil, Usuario, HistorialAjuste,
    Departamento, Categoria, Proveedor, TasaCambio, Marca, ConteoPrioritario, Oferta,
)
from rutas.productos import _precios_computados
from rutas.usuarios import require_admin, require_admin_o_gestionador
from pydantic import BaseModel

router = APIRouter(prefix="/ajustes", tags=["ajustes"])


# ============================================================================
# Schemas
# ============================================================================

class AjusteStockItem(BaseModel):
    producto_id: int
    tipo:        str    # "agregar" | "restar" | "fijar"
    cantidad:    float
    motivo:      str


class AjusteStockLote(BaseModel):
    items: List[AjusteStockItem]


class AjustePrecioItem(BaseModel):
    producto_id: int
    costo_usd:   Optional[float] = None
    margen:      Optional[float] = None


class AjustePrecioLote(BaseModel):
    items: List[AjustePrecioItem]


class AjusteGlobalPrecio(BaseModel):
    filtro_tipo: str            # "todos" | "departamento" | "proveedor" | "pareto"
    filtro_id:   Optional[int] = None
    tipo_ajuste: str            # "costo_pct" | "margen_fijo"
    valor:       float


class AjusteComisionItem(BaseModel):
    producto_id:  int
    comision_pct: float


class AjusteComisionLote(BaseModel):
    items: List[AjusteComisionItem]


class ConteoFisicoItem(BaseModel):
    producto_id:      int
    cantidad_contada: int


class ConteoFisicoLote(BaseModel):
    items: List[ConteoFisicoItem]


# ============================================================================
# Helpers
# ============================================================================

def _tasas(db: Session):
    obj = db.query(TasaCambio).order_by(TasaCambio.id.desc()).first()
    if not obj:
        return 1.0, 1.0
    bcv     = float(obj.tasa or 1)
    binance = float(obj.tasa_binance or bcv)
    return bcv, binance


def _filtrar_productos(db: Session, filtro_tipo: str, filtro_id: Optional[int] = None, categoria_id: Optional[int] = None):
    q = db.query(Producto)
    if filtro_tipo == "departamento" and filtro_id:
        q = q.filter(Producto.departamento_id == filtro_id)
        if categoria_id:
            q = q.filter(Producto.categoria_id == categoria_id)
    elif filtro_tipo == "proveedor" and filtro_id:
        q = q.filter(Producto.proveedor_id == filtro_id)
    elif filtro_tipo == "marca" and filtro_id:
        q = q.filter(Producto.marca_id == filtro_id)
    elif filtro_tipo == "pareto":
        q = q.filter(Producto.es_producto_clave == True)
    return q.all()


def _stock_real(producto: Producto, db: Session) -> int:
    """Para productos con variantes devuelve la suma del stock de variantes."""
    variantes = db.query(VarianteProducto).filter(
        VarianteProducto.producto_id == producto.id
    ).all()
    if variantes:
        return sum(int(v.stock or 0) for v in variantes)
    return int(producto.stock or 0)


def _tiene_variantes(producto_id: int, db: Session) -> int:
    return db.query(VarianteProducto).filter(
        VarianteProducto.producto_id == producto_id
    ).count()


def _guardar_historial(db, usuario, tipo, descripcion, cambios):
    db.add(HistorialAjuste(
        usuario             = usuario or "admin",
        tipo                = tipo,
        descripcion         = descripcion,
        productos_afectados = len(cambios),
        detalle_json        = json.dumps(cambios, ensure_ascii=False),
    ))
    db.commit()


# ============================================================================
# Endpoints — rutas literales antes que cualquier /{id}
# ============================================================================

# ── Catálogo de productos con filtros ────────────────────────────────────────

@router.get("/productos")
def listar_productos_ajuste(
    filtro_tipo:  str           = "todos",
    filtro_id:    Optional[int] = None,
    categoria_id: Optional[int] = None,
    db: Session   = Depends(get_db),
    _: None       = Depends(require_admin_o_gestionador),
):
    productos     = _filtrar_productos(db, filtro_tipo, filtro_id, categoria_id)
    bcv, binance  = _tasas(db)

    deptos_map  = {d.id: d for d in db.query(Departamento).all()}
    provs_map   = {p.id: p for p in db.query(Proveedor).all()}
    marcas_map  = {m.id: m for m in db.query(Marca).all()}
    cats_map    = {c.id: c for c in db.query(Categoria).all()}

    hoy = date.today()
    _ofertas = db.query(Oferta).filter(
        Oferta.activo == True,
        Oferta.fecha_inicio <= hoy,
        or_(Oferta.fecha_fin == None,       Oferta.fecha_fin    >= hoy),
        or_(Oferta.cantidad_limite == None, Oferta.cantidad_usada < Oferta.cantidad_limite),
    ).all()
    ofertas_map = {o.producto_id: o for o in _ofertas}

    result = []
    for p in productos:
        depto           = deptos_map.get(p.departamento_id)
        prov            = provs_map.get(p.proveedor_id)
        marca           = marcas_map.get(p.marca_id) if p.marca_id else None
        cat             = cats_map.get(p.categoria_id) if p.categoria_id else None
        n_variantes     = _tiene_variantes(p.id, db)
        stock_real      = _stock_real(p, db)

        # pricing — resuelve policy sin queries adicionales usando provs_map
        if p.pricing_policy_override:
            pol, at, ap = p.pricing_policy_override, "manual", 0.0
        elif prov and prov.pricing_policy:
            pol = prov.pricing_policy
            at  = getattr(prov, 'ajuste_tipo', 'manual') or 'manual'
            ap  = float(prov.ajuste_divisa_pct or 0.0)
        else:
            pol, at, ap = "MARKET_FACTOR", "manual", 0.0
        precios = _precios_computados(p, bcv, binance, pol, at, ap)

        # oferta activa
        oferta = ofertas_map.get(p.id)
        if oferta:
            precio_oferta = round(precios["precio_base_usd"] * (1 - oferta.valor / 100), 4) \
                            if oferta.tipo_precio == "porcentaje" else oferta.valor
        else:
            precio_oferta = None

        result.append({
            "id":                   p.id,
            "codigo":               p.codigo,
            "nombre":               p.nombre,
            "departamento_id":      p.departamento_id,
            "departamento_nombre":  depto.nombre  if depto  else "—",
            "categoria_id":         p.categoria_id,
            "categoria_nombre":     cat.nombre    if cat    else "—",
            "proveedor_id":         p.proveedor_id,
            "proveedor_nombre":     prov.nombre   if prov   else "—",
            "marca_id":             p.marca_id,
            "marca_nombre":         marca.nombre  if marca  else "—",
            "costo_usd":            float(p.costo_usd    or 0),
            "margen":               float(p.margen        or 0),
            "comision_pct":         float(p.comision_pct  or 0),
            "precio_base_usd":      precios["precio_base_usd"],
            "precio_referencial_usd": precios["precio_referencial_usd"],
            "stock":                stock_real,
            "es_producto_clave":    p.es_producto_clave,
            "es_delicado":          p.es_delicado,
            "tiene_variantes":      n_variantes > 0,
            "variantes_count":      n_variantes,
            "auditado":             bool(p.auditado),
            "auditoria_pendiente":  bool(p.auditoria_pendiente),
            "fecha_auditoria":      p.fecha_auditoria.strftime('%d/%m/%Y') if p.fecha_auditoria else None,
            "conteo_pendiente":     p.conteo_pendiente,
            "diferencia_pendiente": p.diferencia_pendiente,
            "oferta_activa":        oferta is not None,
            "precio_oferta_usd":    precio_oferta,
        })
    return result


# ── Ajuste de stock por lotes ────────────────────────────────────────────────

@router.post("/stock/lote")
def ajuste_stock_lote(
    datos: AjusteStockLote,
    db:    Session           = Depends(get_db),
    _:     None              = Depends(require_admin_o_gestionador),
    x_usuario_nombre: Optional[str] = Header(None),
):
    cambios = []
    for item in datos.items:
        p = db.query(Producto).filter(Producto.id == item.producto_id).first()
        if not p:
            continue
        ant = p.stock
        if item.tipo == "agregar":
            p.stock = p.stock + int(item.cantidad)
        elif item.tipo == "restar":
            p.stock = max(0, p.stock - int(item.cantidad))
        elif item.tipo == "fijar":
            p.stock = int(item.cantidad)
        cambios.append({
            "producto_id":    p.id,
            "nombre":         p.nombre,
            "stock_anterior": ant,
            "stock_nuevo":    p.stock,
            "tipo":           item.tipo,
            "motivo":         item.motivo,
        })
    db.commit()
    _guardar_historial(
        db, x_usuario_nombre, "stock",
        f"Ajuste de stock — {len(cambios)} productos", cambios,
    )
    return {"ok": True, "productos_afectados": len(cambios)}


# ── Ajuste de precios por lotes ──────────────────────────────────────────────

@router.post("/precio/lote")
def ajuste_precio_lote(
    datos: AjustePrecioLote,
    db:    Session           = Depends(get_db),
    _:     None              = Depends(require_admin_o_gestionador),
    x_usuario_nombre: Optional[str] = Header(None),
):
    cambios = []
    for item in datos.items:
        p = db.query(Producto).filter(Producto.id == item.producto_id).first()
        if not p:
            continue
        c_ant, m_ant = p.costo_usd, p.margen
        if item.costo_usd is not None:
            p.costo_usd = round(item.costo_usd, 4)
        if item.margen is not None:
            p.margen = round(item.margen, 6)
        cambios.append({
            "producto_id":    p.id,
            "nombre":         p.nombre,
            "costo_anterior": c_ant,
            "costo_nuevo":    p.costo_usd,
            "margen_anterior":m_ant,
            "margen_nuevo":   p.margen,
        })
    db.commit()
    _guardar_historial(
        db, x_usuario_nombre, "precio",
        f"Ajuste de precios por lotes — {len(cambios)} productos", cambios,
    )
    return {"ok": True, "productos_afectados": len(cambios)}


# ── Ajuste global de precios por filtro ──────────────────────────────────────

@router.post("/precio/global")
def ajuste_precio_global(
    datos: AjusteGlobalPrecio,
    db:    Session             = Depends(get_db),
    _:     None                = Depends(require_admin),
    x_usuario_nombre: Optional[str] = Header(None),
):
    productos = _filtrar_productos(db, datos.filtro_tipo, datos.filtro_id)
    cambios = []
    for p in productos:
        c_ant, m_ant = p.costo_usd, p.margen
        if datos.tipo_ajuste == "costo_pct":
            p.costo_usd = round(float(p.costo_usd or 0) * (1 + datos.valor / 100), 4)
        elif datos.tipo_ajuste == "margen_fijo":
            p.margen = round(datos.valor / 100, 6)
        cambios.append({
            "producto_id":    p.id,
            "nombre":         p.nombre,
            "costo_anterior": c_ant,
            "costo_nuevo":    p.costo_usd,
            "margen_anterior":m_ant,
            "margen_nuevo":   p.margen,
        })
    db.commit()
    desc = f"Ajuste global ({datos.tipo_ajuste}={datos.valor}) — {len(cambios)} productos"
    _guardar_historial(db, x_usuario_nombre, "precio", desc, cambios)
    return {"ok": True, "productos_afectados": len(cambios)}


# ── Ajuste de comisiones por lotes ───────────────────────────────────────────

@router.post("/comisiones/lote")
def ajuste_comisiones_lote(
    datos: AjusteComisionLote,
    db:    Session             = Depends(get_db),
    _:     None                = Depends(require_admin),
    x_usuario_nombre: Optional[str] = Header(None),
):
    cambios = []
    for item in datos.items:
        p = db.query(Producto).filter(Producto.id == item.producto_id).first()
        if not p:
            continue
        ant = p.comision_pct
        p.comision_pct = round(item.comision_pct, 6)
        cambios.append({
            "producto_id":      p.id,
            "nombre":           p.nombre,
            "comision_anterior":ant,
            "comision_nueva":   p.comision_pct,
        })
    db.commit()
    _guardar_historial(
        db, x_usuario_nombre, "comision",
        f"Ajuste de comisión por producto — {len(cambios)} productos", cambios,
    )
    return {"ok": True, "productos_afectados": len(cambios)}


# ── Gestión: mover departamento masivo ───────────────────────────────────────

class MoverProductosLote(BaseModel):
    producto_ids:    List[int]
    departamento_id: Optional[int] = None
    categoria_id:    Optional[int] = None
    proveedor_id:    Optional[int] = None
    marca_id:        Optional[int] = None


@router.post("/productos/mover-departamento")
def mover_departamento_lote(
    datos: MoverProductosLote,
    db:    Session = Depends(get_db),
    _:     None    = Depends(require_admin_o_gestionador),
    x_usuario_nombre: Optional[str] = Header(None),
):
    from fastapi import HTTPException
    deptos_map = {d.id: d for d in db.query(Departamento).all()}
    if datos.departamento_id and datos.departamento_id not in deptos_map:
        raise HTTPException(404, "Departamento no encontrado")

    cambios = []
    for pid in datos.producto_ids:
        p = db.query(Producto).filter(Producto.id == pid).first()
        if not p:
            continue
        cambios.append({"producto_id": p.id, "nombre": p.nombre})
        if datos.departamento_id is not None:
            p.departamento_id = datos.departamento_id
            p.categoria_id    = datos.categoria_id
        if datos.proveedor_id is not None:
            p.proveedor_id = datos.proveedor_id
        if datos.marca_id is not None:
            p.marca_id = datos.marca_id
    db.commit()
    depto_label = deptos_map[datos.departamento_id].nombre if datos.departamento_id else "—"
    _guardar_historial(
        db, x_usuario_nombre, "gestion",
        f"Cambio masivo — {len(cambios)} productos (depto: {depto_label})",
        cambios,
    )
    return {"ok": True, "productos_afectados": len(cambios)}


# ── Gestión: editar nombre ────────────────────────────────────────────────────

class EditarNombreItem(BaseModel):
    producto_id: int
    nombre:      str


@router.post("/productos/editar-nombre")
def editar_nombre_producto(
    datos: EditarNombreItem,
    db:    Session = Depends(get_db),
    _:     None    = Depends(require_admin_o_gestionador),
    x_usuario_nombre: Optional[str] = Header(None),
):
    from fastapi import HTTPException
    p = db.query(Producto).filter(Producto.id == datos.producto_id).first()
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    nombre_anterior = p.nombre
    p.nombre = datos.nombre.strip()
    db.commit()
    _guardar_historial(
        db, x_usuario_nombre, "gestion",
        f"Nombre editado: '{nombre_anterior}' → '{p.nombre}'",
        [{"producto_id": p.id, "nombre_anterior": nombre_anterior,
          "nombre_nuevo": p.nombre}],
    )
    return {"ok": True, "id": p.id, "nombre": p.nombre}


# ── Gestión: crear producto ───────────────────────────────────────────────────

class CrearProductoSchema(BaseModel):
    nombre:          str
    departamento_id: Optional[int] = None
    categoria_id:    Optional[int] = None
    proveedor_id:    Optional[int] = None
    costo_usd:       float = 0.0
    margen:          float = 0.30
    stock:           int   = 0


@router.post("/productos/crear")
def crear_producto_ajuste(
    datos: CrearProductoSchema,
    db:    Session = Depends(get_db),
    _:     None    = Depends(require_admin_o_gestionador),
    x_usuario_nombre: Optional[str] = Header(None),
):
    from fastapi import HTTPException
    if not datos.nombre.strip():
        raise HTTPException(400, "El nombre es obligatorio")
    nuevo = Producto(
        nombre          = datos.nombre.strip(),
        departamento_id = datos.departamento_id,
        categoria_id    = datos.categoria_id,
        proveedor_id    = datos.proveedor_id,
        costo_usd       = datos.costo_usd,
        margen          = datos.margen,
        stock           = datos.stock,
        activo          = True,
    )
    db.add(nuevo)
    db.commit()
    db.refresh(nuevo)
    _guardar_historial(
        db, x_usuario_nombre, "gestion",
        f"Producto creado: '{nuevo.nombre}'",
        [{"producto_id": nuevo.id, "nombre": nuevo.nombre,
          "departamento_id": nuevo.departamento_id,
          "stock_inicial": nuevo.stock}],
    )
    return {"ok": True, "id": nuevo.id, "nombre": nuevo.nombre}


# ── Conteo físico / Auditoría de inventario ──────────────────────────────────

@router.post("/conteo/registrar")
def registrar_conteo(
    datos: ConteoFisicoLote,
    db:    Session           = Depends(get_db),
    _:     None              = Depends(require_admin_o_gestionador),
    x_usuario_nombre: Optional[str] = Header(None),
):
    cambios = []
    for item in datos.items:
        p = db.query(Producto).filter(Producto.id == item.producto_id).first()
        if not p:
            continue
        stock_actual = int(p.stock or 0)
        diferencia   = item.cantidad_contada - stock_actual

        if diferencia >= 0:
            p.stock                = item.cantidad_contada
            p.auditado             = True
            p.fecha_auditoria      = datetime.utcnow()
            p.auditoria_pendiente  = False
            p.conteo_pendiente     = None
            p.diferencia_pendiente = None
            cambios.append({
                "producto_id":    p.id,
                "nombre":         p.nombre,
                "stock_anterior": stock_actual,
                "stock_nuevo":    p.stock,
                "diferencia":     diferencia,
                "estado":         "auditado",
            })
        else:
            p.auditoria_pendiente  = True
            p.conteo_pendiente     = item.cantidad_contada
            p.diferencia_pendiente = diferencia
            p.auditado             = False
            cambios.append({
                "producto_id":   p.id,
                "nombre":        p.nombre,
                "stock_sistema": stock_actual,
                "conteo":        item.cantidad_contada,
                "diferencia":    diferencia,
                "estado":        "pendiente",
            })

    db.commit()
    _guardar_historial(
        db, x_usuario_nombre, "auditoria",
        f"Conteo físico — {len(cambios)} productos", cambios,
    )
    return {"ok": True, "cambios": cambios}


@router.get("/conteo/pendientes")
def listar_pendientes(
    db: Session = Depends(get_db),
    _:  None    = Depends(require_admin_o_gestionador),
):
    pendientes = db.query(Producto).filter(
        Producto.auditoria_pendiente == True
    ).all()
    return [
        {
            "id":                p.id,
            "nombre":            p.nombre,
            "stock":             p.stock,
            "conteo_pendiente":  p.conteo_pendiente,
            "diferencia_pendiente": p.diferencia_pendiente,
            "departamento":      p.departamento_id,
        }
        for p in pendientes
    ]


@router.post("/conteo/autorizar/{producto_id}")
def autorizar_faltante(
    producto_id: int,
    db: Session  = Depends(get_db),
    _:  None     = Depends(require_admin),
    x_usuario_nombre: Optional[str] = Header(None),
):
    from fastapi import HTTPException
    p = db.query(Producto).filter(Producto.id == producto_id).first()
    if not p:
        raise HTTPException(404, "Producto no encontrado")
    if not p.auditoria_pendiente:
        raise HTTPException(400, "No hay auditoría pendiente")

    stock_antes  = int(p.stock or 0)
    diferencia   = int(p.diferencia_pendiente or 0)
    p.stock      = max(0, stock_antes + diferencia)
    p.auditado             = True
    p.fecha_auditoria      = datetime.utcnow()
    p.auditoria_pendiente  = False
    p.conteo_pendiente     = None
    p.diferencia_pendiente = None
    db.commit()

    _guardar_historial(
        db, x_usuario_nombre, "auditoria",
        f"Faltante autorizado: '{p.nombre}' — diferencia {diferencia}",
        [{"producto_id": p.id, "nombre": p.nombre,
          "stock_anterior": stock_antes, "stock_nuevo": p.stock}],
    )
    return {"ok": True, "stock_nuevo": p.stock}


# ── Historial ────────────────────────────────────────────────────────────────

@router.get("/historial")
def listar_historial(
    fecha_desde: Optional[date] = None,
    fecha_hasta: Optional[date] = None,
    tipo:        Optional[str]  = None,
    db:  Session = Depends(get_db),
    _:   None    = Depends(require_admin),
):
    q = db.query(HistorialAjuste).order_by(HistorialAjuste.fecha.desc())
    if fecha_desde:
        q = q.filter(sqlfunc.date(HistorialAjuste.fecha) >= fecha_desde)
    if fecha_hasta:
        q = q.filter(sqlfunc.date(HistorialAjuste.fecha) <= fecha_hasta)
    if tipo:
        q = q.filter(HistorialAjuste.tipo == tipo)
    return [
        {
            "id":                  h.id,
            "fecha":               h.fecha.isoformat(),
            "usuario":             h.usuario,
            "tipo":                h.tipo,
            "descripcion":         h.descripcion,
            "productos_afectados": h.productos_afectados,
        }
        for h in q.all()
    ]


# ── Exportar Excel ───────────────────────────────────────────────────────────

@router.get("/exportar/excel")
def exportar_excel(
    db: Session = Depends(get_db),
    _:  None    = Depends(require_admin),
):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    registros = db.query(HistorialAjuste).order_by(HistorialAjuste.fecha.desc()).all()

    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Historial Ajustes"

    headers = ["Fecha", "Usuario", "Tipo", "Descripción", "Registros afectados"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="1A1A1A")
    header_font = Font(color="FFCC00", bold=True)
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws.column_dimensions["A"].width = 20
    ws.column_dimensions["B"].width = 18
    ws.column_dimensions["C"].width = 12
    ws.column_dimensions["D"].width = 55
    ws.column_dimensions["E"].width = 20

    for h in registros:
        ws.append([
            h.fecha.strftime("%d/%m/%Y %H:%M") if h.fecha else "",
            h.usuario    or "",
            h.tipo       or "",
            h.descripcion or "",
            h.productos_afectados or 0,
        ])

    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": "attachment; filename=historial_ajustes.xlsx"},
    )


# ── Exportar PDF ─────────────────────────────────────────────────────────────

@router.get("/exportar/pdf")
def exportar_pdf(
    db: Session = Depends(get_db),
    _:  None    = Depends(require_admin),
):
    from reportlab.lib.pagesizes import A4, landscape
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer
    from reportlab.lib.styles import getSampleStyleSheet
    from reportlab.lib import colors

    registros = db.query(HistorialAjuste).order_by(HistorialAjuste.fecha.desc()).all()

    output = io.BytesIO()
    doc    = SimpleDocTemplate(output, pagesize=landscape(A4), topMargin=30, bottomMargin=30)
    styles = getSampleStyleSheet()
    elements = [
        Paragraph("Historial de Ajustes — Ferreutil", styles["Title"]),
        Spacer(1, 12),
    ]

    data = [["Fecha", "Usuario", "Tipo", "Descripción", "Afectados"]]
    for h in registros:
        data.append([
            h.fecha.strftime("%d/%m/%Y %H:%M") if h.fecha else "",
            h.usuario    or "",
            h.tipo       or "",
            (h.descripcion or "")[:70],
            str(h.productos_afectados or 0),
        ])

    tbl = Table(data, colWidths=[110, 80, 60, 310, 60])
    tbl.setStyle(TableStyle([
        ("BACKGROUND",    (0, 0), (-1, 0), colors.HexColor("#1A1A1A")),
        ("TEXTCOLOR",     (0, 0), (-1, 0), colors.HexColor("#FFCC00")),
        ("FONTNAME",      (0, 0), (-1, 0), "Helvetica-Bold"),
        ("FONTSIZE",      (0, 0), (-1, -1), 8),
        ("ROWBACKGROUNDS",(0, 1), (-1, -1), [colors.white, colors.HexColor("#F5F5F0")]),
        ("GRID",          (0, 0), (-1, -1), 0.4, colors.HexColor("#DDDDDD")),
        ("VALIGN",        (0, 0), (-1, -1), "MIDDLE"),
        ("LEFTPADDING",   (0, 0), (-1, -1), 5),
        ("RIGHTPADDING",  (0, 0), (-1, -1), 5),
        ("TOPPADDING",    (0, 0), (-1, -1), 4),
        ("BOTTOMPADDING", (0, 0), (-1, -1), 4),
    ]))
    elements.append(tbl)

    doc.build(elements)
    output.seek(0)
    return StreamingResponse(
        output,
        media_type="application/pdf",
        headers={"Content-Disposition": "attachment; filename=historial_ajustes.pdf"},
    )


# ── Trazabilidad por producto ─────────────────────────────────────────────────

@router.get("/trazabilidad")
def trazabilidad_producto(
    producto_id:     Optional[int]  = None,
    departamento_id: Optional[int]  = None,
    categoria_id:    Optional[int]  = None,
    marca_id:        Optional[int]  = None,
    fecha_desde:     Optional[date] = None,
    fecha_hasta:     Optional[date] = None,
    tipo:            Optional[str]  = None,
    db: Session      = Depends(get_db),
    _: None          = Depends(require_admin),
):
    """
    Historial completo de movimientos de un producto:
    compras, ventas, ajustes de stock y devoluciones.
    """
    from models import (
        DetalleVenta, Venta, DetalleRecepcion, RecepcionCompra,
        OrdenCompra, DevolucionCliente, DetalleDevolucionCliente,
        Proveedor,
    )

    ids_productos = []
    if producto_id:
        ids_productos = [producto_id]
    else:
        q = db.query(Producto.id)
        if departamento_id:
            q = q.filter(Producto.departamento_id == departamento_id)
        if categoria_id:
            q = q.filter(Producto.categoria_id == categoria_id)
        if marca_id:
            q = q.filter(Producto.marca_id == marca_id)
        ids_productos = [r[0] for r in q.all()]

    if not ids_productos:
        return []

    movimientos = []

    # ── Ventas ────────────────────────────────────────────────────────────
    q_ventas = db.query(DetalleVenta, Venta).join(
        Venta, DetalleVenta.venta_id == Venta.id
    ).filter(DetalleVenta.producto_id.in_(ids_productos))
    if fecha_desde:
        q_ventas = q_ventas.filter(Venta.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        q_ventas = q_ventas.filter(Venta.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))
    for dv, v in q_ventas.all():
        prod = db.query(Producto).filter(Producto.id == dv.producto_id).first()
        movimientos.append({
            "fecha":          v.fecha.isoformat() if v.fecha else None,
            "tipo":           "venta",
            "producto_id":    dv.producto_id,
            "producto_nombre":prod.nombre if prod else f"ID {dv.producto_id}",
            "cantidad":       -abs(float(dv.cantidad or 0)),
            "referencia":     f"Venta #{v.id}",
            "usuario":        v.usuario,
            "precio_usd":     float(dv.precio_unitario or 0),
            "detalle":        f"Venta #{v.id} — {v.usuario}",
        })

    # ── Compras (recepciones) ─────────────────────────────────────────────
    q_compras = db.query(DetalleRecepcion, RecepcionCompra, OrdenCompra).join(
        RecepcionCompra, DetalleRecepcion.recepcion_id == RecepcionCompra.id
    ).join(
        OrdenCompra, RecepcionCompra.orden_id == OrdenCompra.id
    ).filter(DetalleRecepcion.producto_id.in_(ids_productos))
    if fecha_desde:
        q_compras = q_compras.filter(RecepcionCompra.fecha_recepcion >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        q_compras = q_compras.filter(RecepcionCompra.fecha_recepcion <= datetime.combine(fecha_hasta, datetime.max.time()))
    for dr, rc, oc in q_compras.all():
        prod = db.query(Producto).filter(Producto.id == dr.producto_id).first()
        prov = db.query(Proveedor).filter(Proveedor.id == oc.proveedor_id).first()
        movimientos.append({
            "fecha":          rc.fecha_recepcion.isoformat() if rc.fecha_recepcion else None,
            "tipo":           "compra",
            "producto_id":    dr.producto_id,
            "producto_nombre":prod.nombre if prod else f"ID {dr.producto_id}",
            "cantidad":       abs(float(dr.cantidad_recibida or 0)),
            "referencia":     oc.numero,
            "usuario":        rc.recibido_por,
            "precio_usd":     float(dr.precio_unitario_real_usd or 0),
            "detalle":        f"{oc.numero} — {prov.nombre if prov else ''}",
        })

    # ── Ajustes de stock ──────────────────────────────────────────────────
    q_ajustes = db.query(HistorialAjuste).filter(HistorialAjuste.tipo == "stock")
    if fecha_desde:
        q_ajustes = q_ajustes.filter(HistorialAjuste.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        q_ajustes = q_ajustes.filter(HistorialAjuste.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))
    for aj in q_ajustes.all():
        try:
            detalles = json.loads(aj.detalle_json or "[]")
            for d in detalles:
                if d.get("producto_id") in ids_productos:
                    diff = (d.get("stock_nuevo", 0) or 0) - (d.get("stock_anterior", 0) or 0)
                    movimientos.append({
                        "fecha":          aj.fecha.isoformat() if aj.fecha else None,
                        "tipo":           "ajuste",
                        "producto_id":    d["producto_id"],
                        "producto_nombre":d.get("nombre", ""),
                        "cantidad":       diff,
                        "referencia":     f"Ajuste #{aj.id}",
                        "usuario":        aj.usuario,
                        "precio_usd":     None,
                        "detalle":        d.get("motivo", aj.descripcion),
                    })
        except Exception:
            pass

    # ── Devoluciones de clientes ───────────────────────────────────────────
    q_devs = db.query(DetalleDevolucionCliente, DevolucionCliente).join(
        DevolucionCliente, DetalleDevolucionCliente.devolucion_id == DevolucionCliente.id
    ).filter(DetalleDevolucionCliente.producto_id.in_(ids_productos))
    if fecha_desde:
        q_devs = q_devs.filter(DevolucionCliente.fecha >= datetime.combine(fecha_desde, datetime.min.time()))
    if fecha_hasta:
        q_devs = q_devs.filter(DevolucionCliente.fecha <= datetime.combine(fecha_hasta, datetime.max.time()))
    for dd, dev in q_devs.all():
        movimientos.append({
            "fecha":          dev.fecha.isoformat() if dev.fecha else None,
            "tipo":           "devolucion",
            "producto_id":    dd.producto_id,
            "producto_nombre":dd.nombre_producto,
            "cantidad":       abs(float(dd.cantidad or 0)) if dd.vuelve_inventario else 0,
            "referencia":     f"Dev #{dev.id}",
            "usuario":        dev.usuario,
            "precio_usd":     float(dd.precio_unitario or 0),
            "detalle":        f"Devolución — {dev.motivo}",
        })

    movimientos.sort(key=lambda x: x["fecha"] or "", reverse=True)
    if tipo:
        movimientos = [m for m in movimientos if m["tipo"] == tipo]
    return movimientos


# ─────────────────────────────────────────────────────────────────────────────
# CONTEO PRIORITARIO
# ─────────────────────────────────────────────────────────────────────────────

class EnvioConteoItem(BaseModel):
    producto_id: int
    nota: Optional[str] = None
    prioridad: str = "manual"

class EnvioConteoPayload(BaseModel):
    items: list[EnvioConteoItem]
    enviado_por: str

@router.post("/conteo-prioritario/enviar")
def enviar_a_conteo_prioritario(payload: EnvioConteoPayload, db: Session = Depends(get_db),
                                 _: None = Depends(require_admin)):
    """Admin envía productos a la cola de conteo prioritario.
    Si un producto ya está en cola pendiente, no lo duplica."""
    agregados = 0
    duplicados = 0
    for item in payload.items:
        ya_en_cola = db.query(ConteoPrioritario).filter(
            ConteoPrioritario.producto_id == item.producto_id,
            ConteoPrioritario.estado == "pendiente",
        ).first()
        if ya_en_cola:
            duplicados += 1
            continue
        c = ConteoPrioritario(
            producto_id = item.producto_id,
            enviado_por = payload.enviado_por,
            nota        = item.nota,
            prioridad   = item.prioridad,
        )
        db.add(c)
        agregados += 1
    db.commit()
    return {"agregados": agregados, "duplicados": duplicados}

@router.get("/conteo-prioritario/pendientes")
def listar_cola_conteo(db: Session = Depends(get_db),
                       _: None = Depends(require_admin_o_gestionador)):
    """Lista productos en cola, ordenados por prioridad."""
    ORDEN = {"delicado": 1, "compra_entrante": 2, "manual": 3, "top_vendidos": 4}
    items = db.query(ConteoPrioritario).filter(
        ConteoPrioritario.estado == "pendiente"
    ).all()
    resultado = []
    for c in items:
        p = db.query(Producto).filter(Producto.id == c.producto_id).first()
        if not p:
            continue
        resultado.append({
            "id":             c.id,
            "producto_id":    p.id,
            "nombre":         p.nombre,
            "codigo":         p.codigo,
            "stock_actual":   p.stock,
            "enviado_por":    c.enviado_por,
            "fecha_envio":    c.fecha_envio.isoformat() if c.fecha_envio else None,
            "nota":           c.nota,
            "prioridad":      c.prioridad,
            "es_delicado":    p.es_delicado,
            "es_pareto":      p.es_producto_clave,
        })
    resultado.sort(key=lambda x: ORDEN.get(x["prioridad"], 99))
    return resultado

class ConteoRealizado(BaseModel):
    stock_real: int

@router.post("/conteo-prioritario/{conteo_id}/contar")
def registrar_conteo_prioritario(conteo_id: int, datos: ConteoRealizado,
                                  db: Session = Depends(get_db),
                                  _: None = Depends(require_admin_o_gestionador),
                                  x_usuario: str = Header(None)):
    """Gestionador registra el conteo. Queda pendiente de aprobación admin
    si hay diferencia. Si no hay diferencia, se cierra automáticamente."""
    c = db.query(ConteoPrioritario).filter(ConteoPrioritario.id == conteo_id).first()
    if not c or c.estado != "pendiente":
        raise HTTPException(status_code=404, detail="Conteo no encontrado o ya cerrado")
    p = db.query(Producto).filter(Producto.id == c.producto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    c.stock_sistema   = p.stock
    c.stock_real      = datos.stock_real
    c.diferencia      = datos.stock_real - p.stock
    c.fecha_conteo    = datetime.utcnow()
    c.contado_por     = x_usuario or "gestionador"

    if c.diferencia == 0:
        c.estado          = "contado"
        c.aprobado_admin  = True
        c.fecha_aprobacion= datetime.utcnow()
        p.auditado        = True
        p.fecha_auditoria = datetime.utcnow()
    else:
        c.estado = "contado"

    db.commit()
    return {"ok": True, "diferencia": c.diferencia,
            "requiere_aprobacion": c.diferencia != 0}

@router.get("/conteo-prioritario/discrepancias")
def listar_discrepancias_pendientes(db: Session = Depends(get_db),
                                     _: None = Depends(require_admin)):
    """Lista conteos con discrepancia pendientes de aprobación admin."""
    items = db.query(ConteoPrioritario).filter(
        ConteoPrioritario.estado == "contado",
        ConteoPrioritario.aprobado_admin.is_(None),
    ).all()
    resultado = []
    for c in items:
        p = db.query(Producto).filter(Producto.id == c.producto_id).first()
        if not p: continue
        resultado.append({
            "id":             c.id,
            "producto_id":    p.id,
            "nombre":         p.nombre,
            "stock_sistema":  c.stock_sistema,
            "stock_real":     c.stock_real,
            "diferencia":     c.diferencia,
            "contado_por":    c.contado_por,
            "fecha_conteo":   c.fecha_conteo.isoformat() if c.fecha_conteo else None,
            "nota":           c.nota,
        })
    return resultado

@router.post("/conteo-prioritario/{conteo_id}/aprobar")
def aprobar_discrepancia(conteo_id: int, aprobar: bool = True,
                          db: Session = Depends(get_db),
                          _: None = Depends(require_admin)):
    """Admin aprueba o rechaza la discrepancia. Si aprueba, ajusta el stock."""
    c = db.query(ConteoPrioritario).filter(ConteoPrioritario.id == conteo_id).first()
    if not c:
        raise HTTPException(status_code=404, detail="Conteo no encontrado")
    p = db.query(Producto).filter(Producto.id == c.producto_id).first()
    if not p:
        raise HTTPException(status_code=404, detail="Producto no encontrado")

    c.aprobado_admin   = aprobar
    c.fecha_aprobacion = datetime.utcnow()

    if aprobar:
        p.stock = c.stock_real
        p.auditado = True
        p.fecha_auditoria = datetime.utcnow()
        h = HistorialAjuste(
            producto_id     = p.id,
            tipo            = "conteo_prioritario",
            cantidad_antes  = c.stock_sistema,
            cantidad_despues= c.stock_real,
            diferencia      = c.diferencia,
            motivo          = f"Conteo prioritario aprobado. {c.nota or ''}",
            usuario         = "admin",
            fecha           = datetime.utcnow(),
        )
        db.add(h)

    db.commit()
    return {"ok": True}
